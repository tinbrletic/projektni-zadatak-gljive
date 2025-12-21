from openpyxl import load_workbook
import peptides
import mysql.connector


# Computes physicochemical properties for a single peptide
def physicochemical(seq):
  peptide = peptides.Peptide(seq)
  return [
      peptide.cruciani_properties()[0], #0
      peptide.cruciani_properties()[1], #1
      peptide.cruciani_properties()[2], #2
      peptide.instability_index(), #3
      peptide.boman(), #4
      peptide.hydrophobicity("KyteDoolittle"), #5
      peptide.hydrophobicity("HoppWoods"), #6
      #peptide.hydrophobicity("Cornette"), #7
      None,
      peptide.hydrophobicity("Eisenberg"), #8
      peptide.hydrophobicity("Roseman"), #9
      peptide.hydrophobicity("Janin"), #10
      peptide.hydrophobicity("Engelman"), #11
      peptide.hydrophobic_moment(angle=100, window=min(len(seq), 11)), #12
      peptide.aliphatic_index(), #13
      peptide.isoelectric_point("Lehninger"), #14
      peptide.charge(pKscale="Lehninger"), #15
  ]
# print("Physicochemical for ACDC:", physicochemical("ACDC"))


# Computes compositional properties
# Corrected categorization based on standard biochemical properties (Wikipedia)
def compositional(seq):
  groups = (
      ('A', 'C', 'G', 'S', 'T'),                                  #0 Tiny 
      ('A', 'C', 'D', 'G', 'N', 'P', 'S', 'T', 'V'),              #1 Small
      ('A', 'I', 'L', 'M', 'V'),                                  #2 Aliphatic (added M, per Wikipedia)
      ('F', 'W', 'Y', 'H'),                                       #3 Aromatic (removed H, per Wikipedia)
      ('A', 'C', 'F', 'G', 'I', 'L', 'M', 'P', 'V', 'W', 'Y'),    #4 Non-polar (removed H)
      ('D', 'E', 'H', 'K', 'N', 'Q', 'R', 'S', 'T'),              #5 Polar
      ('D', 'E', 'H', 'K', 'R'),                                  #6 Charged
      ('H', 'K', 'R'),                                            #7 Basic
      ('D', 'E')                                                  #8 Acidic
  )

  properties = []
  for group in groups:
    count = 0
    for amino in group:
      count += seq.count(amino)
    properties.append(count)
    properties.append(count / len(seq))
  return properties

# print("Compositional for ACDC:", compositional("ACDC"))


excel_file = 'ao2c02425_si_002.xlsx'
wb = load_workbook(excel_file)
ws = wb.active

connection = mysql.connector.connect(
    host="127.0.0.1", # MySQL adresa servera (localhost)
    user="root",  # MySQL korisničko ime
    password="Databejs567!",  # MySQL lozinka
    database="peptide-dataset"  # Naziv baze podataka
)
cursor = connection.cursor()

# cursor.execute("""
# CREATE TABLE IF NOT EXISTS peptides (
#     id INT AUTO_INCREMENT PRIMARY KEY,
#     peptide_seq VARCHAR(255),
#     peptide_len INT,
#     hydrophobic_index DOUBLE,
#     synthesis_flag VARCHAR(255)
# )
# """)

print(peptides.tables.HYDROPHOBICITY.keys())

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=3):
  if row[0].value is not None:
    pep = peptides.Peptide(row[0].value)
  else:
    print("Prazna ćelija pronađena u stupcu za sekvencu peptida.")
    continue

  print(row[0].value, pep)
  phy = physicochemical(row[0].value)
  comp = compositional(row[0].value)

  # 4. Unos podataka u bazu

  cursor.execute("""INSERT INTO peptides (peptide_seq, peptide_len, synthesis_flag, hydrophobic_janin, hydrophobic_engleman, hydrophobic_moment, aliphatic_index, isoelectric_point, charge, tiny_group, small_group, aliphatic_group, aromatic_group, `non-polar_group`, polar_group, charged_group, basic_group, acidic_group, cruciani_prp1, cruciani_prp2, cruciani_prp3, instability_index, boman, `hydrophobic_kyte-doolittle`, `hydrophobic_hopp-woods`, `hydrophobic_cornette`, `hydrophobic_eisenberg`, `hydrophobic_roseman`) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", (row[0].value, row[2].value, ('TRUE' if row[1].value == 1 else 'FALSE'), phy[10], phy[11], phy[12], phy[13], phy[14], phy[15], comp[0], comp[1], comp[2], comp[3], comp[4], comp[5], comp[6], comp[7], comp[8], phy[0], phy[1], phy[2], phy[3], phy[4], phy[5], phy[6], phy[7], phy[8], phy[9]))

# 5. Spremanje promjena i zatvaranje konekcije
connection.commit()
connection.close()
wb.close()

## task 1. random forest feature selection
## task 2. sequntion feature selection